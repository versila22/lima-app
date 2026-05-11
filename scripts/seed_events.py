"""
Seed events from Excel calendar into LIMA via API.
Usage: python seed_events.py
"""
import json
import urllib.request
import urllib.error
import http.cookiejar
from datetime import date, datetime

import openpyxl

EXCEL_PATH = r"C:\Users\jerom\Downloads\calendrierlima2526_V12.xlsx"
API_BASE = "https://api-production-e15b.up.railway.app"
ADMIN_EMAIL = "admin@lima-impro.fr"
ADMIN_PASSWORD = "Admin1234!"

# ── Helpers ──────────────────────────────────────────────────────────────────

def api(jar, method, path, body=None):
    url = API_BASE + path
    data = json.dumps(body).encode() if body else None
    headers = {"Content-Type": "application/json", "Accept": "application/json"}
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(jar))
    try:
        with opener.open(req) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"{method} {path} → {e.code}: {e.read().decode()}")


# ── Extract events from Excel ─────────────────────────────────────────────────

HOLIDAYS = {
    'toussaint', 'armistice', 'assomption', "jour de l'an", 'saint-sylvestre',
    'noël', 'noel', 'paques', 'pâques', 'ascension', 'pentecôte', 'pentecote',
    'fête nationale', 'fete nationale', 'fête du travail', 'fete du travail',
    'fête de la victoire', 'fete de la victoire', 'passage', 'germoir',
    'festival champ', 'digital village', 'informations', 'réveil', 'reveil',
}


def is_noise(t):
    tl = t.lower()
    return any(h in tl for h in HOLIDAYS)


def classify(text):
    """Return (event_type, is_away) or None to skip."""
    if not text:
        return None
    t = str(text).strip()
    if not t:
        return None
    if is_noise(t):
        return None
    tl = t.lower()
    if 'entrainement loisirs' in tl or 'entraînement loisirs' in tl:
        return ('training_leisure', False)
    if 'entrainement spectacle' in tl or 'entraînement spectacle' in tl:
        return ('training_show', False)
    if tl.rstrip('-').strip() == 'welsh':
        return None  # Welsh handled separately from Recap
    if 'formation' in tl:
        return ('formation', False)
    if tl.strip() == 'ag':
        return ('ag', False)
    if 'dépla' in tl or 'depla' in tl:
        # Extract city from "Dépla CITY (opponent)"
        return ('match', True)
    if any(x in tl for x in ['spectacle loisirs', 'spectacle chabrol', 'cabaret', 'giffard', 'hce']):
        return ('cabaret', False)
    if any(x in tl for x in ['match', 'chabrol', 'monplaisir', 'balise', 'muchachos',
                               'improlokos', 'cito', 'patounet', 'carré des arts', 'catch']):
        return ('match', False)
    if 'spectacle' in tl:
        return ('cabaret', False)
    return ('other', False)


def extract_calendar_events():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Calendrier scolaire 2025-2026']
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    SECTIONS = [
        {
            'row_start': 2, 'row_end': 34,
            'months': [(0, 8, 2025), (4, 9, 2025), (8, 10, 2025),
                       (12, 11, 2025), (16, 12, 2025), (20, 1, 2026)],
        },
        {
            'row_start': 37, 'row_end': 71,
            'months': [(0, 2, 2026), (4, 3, 2026), (8, 4, 2026),
                       (12, 5, 2026), (16, 6, 2026), (20, 7, 2026)],
        },
    ]

    events = {}  # (date, normalized_title) → event dict

    for section in SECTIONS:
        for row_idx in range(section['row_start'], section['row_end']):
            row = all_rows[row_idx]
            for col_start, month, year in section['months']:
                day_val = row[col_start]
                if day_val is None:
                    # Infer day from row position
                    day_val = row_idx - section['row_start'] + 1
                event_text = row[col_start + 2] if col_start + 2 < len(row) else None
                if not isinstance(day_val, int) or not event_text:
                    continue
                res = classify(str(event_text))
                if res is None:
                    continue
                try:
                    d = date(year, month, day_val)
                except ValueError:
                    continue
                event_type, is_away = res
                title = str(event_text).strip()
                key = (d, title.lower()[:30])
                if key not in events:
                    events[key] = {
                        'date': d,
                        'title': title,
                        'event_type': event_type,
                        'is_away': is_away,
                        'notes': None,
                    }

    return list(events.values())


def extract_recap_events():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Recap']
    rows = list(ws.iter_rows(min_row=4, values_only=True))

    WELSH_DATES = [
        date(2025, 9, 19), date(2025, 10, 10), date(2025, 11, 7),
        date(2025, 12, 5),  date(2026, 1, 30),  date(2026, 3, 6),
        date(2026, 4, 10),  date(2026, 5, 29),
    ]
    events = []

    # Welsh recurring show
    for d in WELSH_DATES:
        events.append({
            'date': d,
            'title': 'Welsh',
            'event_type': 'welsh',
            'is_away': False,
            'notes': 'Bar Le Welsh - Angers · 4 joueurs · Match/Cab',
        })

    # GIFFARD recurring cabaret
    giffard_dates_str = '16/10 - 11/12 - 19/03 - 29/04 - 22/05 - 18/06'
    for ds in giffard_dates_str.split(' - '):
        ds = ds.strip().rstrip('?')
        try:
            day, month = int(ds[:2]), int(ds[3:5])
            year = 2026 if month <= 6 else 2025
            events.append({
                'date': date(year, month, day),
                'title': 'Cabaret Giffard',
                'event_type': 'cabaret',
                'is_away': False,
                'notes': 'Espace Menthe Pastille - Avrillé · 5 joueurs',
            })
        except Exception:
            pass

    # Individual matches from Recap rows
    RECAP_MATCHES = [
        {'title': 'Match La Balise', 'date': date(2025, 9, 27),  'notes': 'Salle C. Chabrol - Angers · 5 joueurs · La Balise (Limoges)', 'is_away': False},
        {'title': 'Match Muchachos', 'date': date(2025, 10, 18), 'notes': 'MPT Monplaisir - Angers · 5 joueurs · Muchachos (Niort)', 'is_away': False},
        {'title': 'Match Improlokos','date': date(2025, 11, 22), 'notes': 'MPT Monplaisir - Angers · 5 joueurs · Improlokos (Blois)', 'is_away': False},
        {'title': 'Catch Patounets', 'date': date(2025, 11, 28), 'notes': 'MQ St Aubin · 2 joueurs · Invitation Patounets', 'is_away': False},
        {'title': 'Match TIC',       'date': date(2025, 12, 13), 'notes': 'MPT Monplaisir - Angers · 5 joueurs · TIC (Tours)', 'is_away': False},
        {'title': 'Match CITO',      'date': date(2026, 1, 17),  'notes': 'MPT Monplaisir - Angers · 5 joueurs · CITO (Nantes)', 'is_away': False},
    ]
    for m in RECAP_MATCHES:
        events.append({**m, 'event_type': 'match'})

    return events


def merge_events(cal_events, recap_events):
    """Merge, deduplicate Recap matches over calendar matches on same date/type."""
    # Recap events take priority for match/cabaret/welsh on same date
    priority_keys = set()
    for e in recap_events:
        priority_keys.add((e['date'], e['event_type']))

    result = list(recap_events)
    for e in cal_events:
        key = (e['date'], e['event_type'])
        if key in priority_keys and e['event_type'] in ('match', 'cabaret', 'welsh'):
            continue  # Recap version already included
        result.append(e)

    result.sort(key=lambda e: e['date'])
    return result


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Extracting events from Excel...")
    cal_events = extract_calendar_events()
    recap_events = extract_recap_events()
    all_events = merge_events(cal_events, recap_events)

    print(f"  Calendar: {len(cal_events)} events")
    print(f"  Recap:    {len(recap_events)} events")
    print(f"  Merged:   {len(all_events)} events")

    print("\nLogging in as admin...")
    jar = http.cookiejar.CookieJar()
    api(jar, "POST", "/auth/login", {"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
    print("  Logged in OK")

    print("\nFetching current season...")
    seasons = api(jar, "GET", "/seasons")
    current = next((s for s in seasons if s.get("is_current")), None)
    if not current:
        raise RuntimeError("No current season found!")
    season_id = current["id"]
    print(f"  Season: {current['name']} ({season_id})")

    print(f"\nInserting {len(all_events)} events...")
    ok = 0
    errors = []
    for e in all_events:
        payload = {
            "season_id": season_id,
            "title": e["title"],
            "event_type": e["event_type"],
            "start_at": f"{e['date']}T20:00:00",
            "is_away": e["is_away"],
            "visibility": "all",
        }
        if e.get("notes"):
            payload["notes"] = e["notes"]
        try:
            api(jar, "POST", "/events", payload)
            ok += 1
            print(f"  OK {e['date']} [{e['event_type']:18}] {e['title'][:50]}")
        except Exception as exc:
            errors.append(f"  ERR {e['date']} {e['title']}: {exc}")
            print(errors[-1])

    print(f"\nDone: {ok} created, {len(errors)} errors.")
    if errors:
        print("\nErrors:")
        for err in errors:
            print(err)


if __name__ == "__main__":
    main()
