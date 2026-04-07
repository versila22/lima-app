"""
Import service — CSV HelloAsso (members) and Excel calendar import.

HelloAsso exports two CSV files:
  - adherents.csv : personal info + membership fee
  - joueurs.csv   : player info + player fee + player_status

Calendar Excel:
  - Each row is an event with at minimum a date and title column.
"""

import io
import logging
import secrets
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Tuple
from uuid import UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.models.event import Event
from app.models.member import Member
from app.models.member_season import MemberSeason
from app.models.season import Season
from app.models.venue import Venue
from app.schemas.event import CalendarImportReport
from app.schemas.member import ImportMemberReport, MemberSummary
from app.services.email_service import send_activation_email

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Keywords → event_type mapping (case-insensitive substring match)
EVENT_TYPE_KEYWORDS: List[Tuple[str, str]] = [
    ("match", "match"),
    ("cabaret", "cabaret"),
    ("welsh", "welsh"),
    ("giffard", "training_show"),
    ("entraîn", "training_leisure"),
    ("entrainement", "training_leisure"),
    ("formation", "formation"),
    ("ag ", "ag"),
    ("assemblée générale", "ag"),
    ("assemblée", "ag"),
]

# HelloAsso player_status deduced from player_fee range (fallback only)
PLAYER_FEE_STATUS: List[Tuple[float, float, str]] = [
    (150.0, 180.0, "M"),  # Match ~160€
    (0.01, 100.0, "L"),   # Loisir ~75€ (anything below match)
]

# "Groupe de Jeu" column → player_status mapping (takes priority over fee)
GROUPE_JEU_STATUS: dict = {
    "match": "M",
    "cabaret": "C",
    "loisir": "L",
}


# ---------------------------------------------------------------------------
# CSV HelloAsso import
# ---------------------------------------------------------------------------

def _parse_decimal(value: str) -> Optional[Decimal]:
    """Convert a string to Decimal, returning None on failure."""
    try:
        return Decimal(value.replace(",", ".").strip())
    except (InvalidOperation, AttributeError):
        return None


def _normalize_name(value: str) -> str:
    return value.strip().title()


def _normalize_email(value: str) -> str:
    return value.strip().lower()


def _deduce_player_status(player_fee: Optional[Decimal], groupe_jeu: Optional[str] = None) -> str:
    """Deduce player_status from Groupe de Jeu field (priority) or player fee (fallback)."""
    # Priority: explicit group from CSV
    if groupe_jeu:
        status = GROUPE_JEU_STATUS.get(groupe_jeu.strip().lower())
        if status:
            return status
    # Fallback: fee-based detection
    if player_fee is None:
        return "A"
    fee_f = float(player_fee)
    for low, high, status in PLAYER_FEE_STATUS:
        if low <= fee_f <= high:
            return status
    return "A"


async def import_csv_helloasso(
    db: AsyncSession,
    adherents_bytes: bytes,
    joueurs_bytes: bytes,
    season_id: UUID,
) -> ImportMemberReport:
    """
    Parse the two HelloAsso CSV exports and upsert members + member_seasons.

    Args:
        db: Async database session.
        adherents_bytes: Raw bytes of the adherents CSV file.
        joueurs_bytes: Raw bytes of the joueurs CSV file.
        season_id: Target season UUID.

    Returns:
        ImportMemberReport with counts and error messages.
    """
    import csv

    report = ImportMemberReport()

    # --- Parse adherents CSV ---
    adherents: Dict[str, dict] = {}
    try:
        text = adherents_bytes.decode("utf-8-sig")  # Handle BOM
        reader = csv.DictReader(io.StringIO(text), delimiter=";")
        for row in reader:
            email = _normalize_email(row.get("Email", "") or row.get("email", ""))
            if not email:
                continue
            adherents[email] = {
                "email": email,
                "first_name": _normalize_name(
                    row.get("Prénom", "") or row.get("Prenom", "") or row.get("first_name", "")
                ),
                "last_name": _normalize_name(
                    row.get("Nom", "") or row.get("last_name", "")
                ),
                "phone": (row.get("Téléphone", "") or row.get("telephone", "") or "").strip() or None,
                "membership_fee": _parse_decimal(
                    row.get("Montant", "") or row.get("montant", "") or "0"
                ),
                "helloasso_ref": (
                    row.get("N° commande", "") or row.get("ref", "") or ""
                ).strip() or None,
                "address": (row.get("Adresse", "") or "").strip() or None,
                "postal_code": (row.get("Code postal", "") or "").strip() or None,
                "city": (row.get("Ville", "") or "").strip() or None,
            }
    except Exception as exc:
        logger.exception("Erreur parsing adherents CSV")
        report.errors.append(f"Erreur lecture adhérents CSV: {exc}")
        return report

    # --- Parse joueurs CSV ---
    joueurs: Dict[str, dict] = {}
    try:
        text = joueurs_bytes.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text), delimiter=";")
        for row in reader:
            # joueurs CSV uses "Email" column (the player's own email)
            email = _normalize_email(
                row.get("Email", "") or row.get("email", "") or row.get("Email payeur", "") or ""
            )
            if not email:
                continue
            joueurs[email] = {
                "player_fee": _parse_decimal(
                    row.get("Montant tarif", "") or row.get("Montant", "") or row.get("montant", "") or "0"
                ),
                "helloasso_ref": (
                    row.get("Référence commande", "") or row.get("N° commande", "") or row.get("ref", "") or ""
                ).strip() or None,
                "groupe_jeu": (
                    row.get("Groupe de Jeu", "") or row.get("groupe_jeu", "") or ""
                ).strip() or None,
            }
    except Exception as exc:
        logger.exception("Erreur parsing joueurs CSV")
        report.errors.append(f"Erreur lecture joueurs CSV: {exc}")
        return report

    # Verify season exists
    season_result = await db.execute(select(Season).where(Season.id == season_id))
    season = season_result.scalar_one_or_none()
    if season is None:
        report.errors.append(f"Saison {season_id} introuvable")
        return report

    # --- Merge and upsert ---
    all_emails = set(adherents.keys()) | set(joueurs.keys())
    for email in all_emails:
        adh = adherents.get(email, {})
        jou = joueurs.get(email, {})

        # Merge data
        first_name = adh.get("first_name") or ""
        last_name = adh.get("last_name") or ""
        if not first_name or not last_name:
            report.errors.append(f"Données incomplètes pour {email}")
            continue

        player_fee = jou.get("player_fee")
        membership_fee = adh.get("membership_fee")
        groupe_jeu = jou.get("groupe_jeu")
        player_status = _deduce_player_status(player_fee, groupe_jeu)
        helloasso_ref = jou.get("helloasso_ref") or adh.get("helloasso_ref")

        try:
            # Upsert member
            existing = await db.execute(
                select(Member).where(Member.email == email)
            )
            member = existing.scalar_one_or_none()
            is_new_member = member is None

            if member is None:
                member = Member(
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    phone=adh.get("phone"),
                    address=adh.get("address"),
                    postal_code=adh.get("postal_code"),
                    city=adh.get("city"),
                    activation_token=secrets.token_urlsafe(32),
                    activation_expires_at=datetime.now(timezone.utc) + timedelta(days=7),
                )
                db.add(member)
                await db.flush()  # Get ID
                report.created += 1
            else:
                # Only update personal info if not already set
                member.first_name = first_name or member.first_name
                member.last_name = last_name or member.last_name
                if adh.get("phone") and not member.phone:
                    member.phone = adh["phone"]
                if adh.get("address") and not member.address:
                    member.address = adh["address"]
                report.updated += 1

            # Upsert member_season
            ms_result = await db.execute(
                select(MemberSeason).where(
                    MemberSeason.member_id == member.id,
                    MemberSeason.season_id == season_id,
                )
            )
            ms = ms_result.scalar_one_or_none()
            if ms is None:
                ms = MemberSeason(
                    member_id=member.id,
                    season_id=season_id,
                    player_status=player_status,
                    membership_fee=membership_fee,
                    player_fee=player_fee,
                    helloasso_ref=helloasso_ref,
                )
                db.add(ms)
            else:
                ms.player_status = player_status
                ms.membership_fee = membership_fee
                ms.player_fee = player_fee
                if helloasso_ref:
                    ms.helloasso_ref = helloasso_ref

            if is_new_member and settings.SMTP_HOST and member.activation_token:
                await send_activation_email(
                    to=member.email,
                    first_name=member.first_name,
                    token=member.activation_token,
                    base_url=settings.FRONTEND_URL,
                )

            report.members.append(
                MemberSummary(
                    id=member.id,
                    email=member.email,
                    first_name=member.first_name,
                    last_name=member.last_name,
                    app_role=member.app_role,
                    is_active=member.is_active,
                    player_status=player_status,
                )
            )

        except Exception as exc:
            logger.exception(f"Erreur import membre {email}")
            report.errors.append(f"Erreur membre {email}: {exc}")

    await db.flush()
    return report


# ---------------------------------------------------------------------------
# Excel calendar import
# ---------------------------------------------------------------------------

def _detect_event_type(title: str) -> str:
    """Detect event type from title using keyword matching."""
    title_lower = title.lower()
    for keyword, event_type in EVENT_TYPE_KEYWORDS:
        if keyword in title_lower:
            return event_type
    return "other"


def _try_parse_datetime(value) -> Optional[datetime]:
    """Try to parse a cell value as a datetime."""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(value.strip(), fmt)
                return dt.replace(tzinfo=timezone.utc)
            except ValueError:
                pass
    return None


async def import_excel_calendar(
    db: AsyncSession,
    excel_bytes: bytes,
    season_id: UUID,
) -> CalendarImportReport:
    """
    Parse an Excel calendar file and create events in the given season.

    Expected columns (flexible, case-insensitive):
      - date / Date / start_at
      - title / Titre / Description
      - end_date / end / fin (optional)
      - venue / lieu / Lieu (optional — matched by name)
      - notes / Notes (optional)

    Args:
        db: Async database session.
        excel_bytes: Raw bytes of the Excel file.
        season_id: Target season UUID.

    Returns:
        CalendarImportReport with counts.
    """
    import openpyxl

    report = CalendarImportReport()

    # Verify season
    season_result = await db.execute(select(Season).where(Season.id == season_id))
    season = season_result.scalar_one_or_none()
    if season is None:
        report.errors.append(f"Saison {season_id} introuvable")
        return report

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(excel_bytes), read_only=True, data_only=True
        )
        ws = wb.active
    except Exception as exc:
        report.errors.append(f"Impossible d'ouvrir le fichier Excel: {exc}")
        return report

    # Detect header row
    headers: Dict[str, int] = {}
    header_row_idx: Optional[int] = None
    HEADER_ALIASES = {
        "date": ["date", "start_at", "début", "debut"],
        "title": ["title", "titre", "description", "événement", "evenement"],
        "end": ["end", "end_date", "fin"],
        "venue": ["venue", "lieu", "salle"],
        "notes": ["notes", "note", "commentaire"],
    }

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > 10:
            break
        row_strs = [str(c).lower().strip() if c else "" for c in row]
        for canonical, aliases in HEADER_ALIASES.items():
            for col_idx, cell_str in enumerate(row_strs):
                if cell_str in aliases and canonical not in headers:
                    headers[canonical] = col_idx
        if "date" in headers and "title" in headers:
            header_row_idx = row_idx
            break

    if header_row_idx is None or "date" not in headers or "title" not in headers:
        report.errors.append(
            "Colonnes 'date' et 'title' introuvables dans le fichier Excel"
        )
        return report

    # Cache venues by name (lowercase)
    venues_result = await db.execute(select(Venue))
    venues_by_name: Dict[str, UUID] = {
        v.name.lower(): v.id for v in venues_result.scalars().all()
    }

    # Existing events: deduplicate by (season_id, start_at, title)
    existing_events_result = await db.execute(
        select(Event.title, Event.start_at).where(Event.season_id == season_id)
    )
    existing_keys = {
        (row.title.lower(), row.start_at.date() if row.start_at else None)
        for row in existing_events_result.all()
    }

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx <= header_row_idx:
            continue

        try:
            cells = list(row)

            date_val = cells[headers["date"]] if headers["date"] < len(cells) else None
            title_val = cells[headers["title"]] if headers["title"] < len(cells) else None

            if not date_val or not title_val:
                continue

            start_at = _try_parse_datetime(date_val)
            title = str(title_val).strip()

            if not start_at or not title:
                continue

            # Dedup check
            dedup_key = (title.lower(), start_at.date())
            if dedup_key in existing_keys:
                report.skipped += 1
                continue

            # Optional fields
            end_at = None
            if "end" in headers and headers["end"] < len(cells):
                end_at = _try_parse_datetime(cells[headers["end"]])

            venue_id = None
            if "venue" in headers and headers["venue"] < len(cells):
                venue_name = str(cells[headers["venue"]] or "").strip().lower()
                venue_id = venues_by_name.get(venue_name)

            notes = None
            if "notes" in headers and headers["notes"] < len(cells):
                notes_val = cells[headers["notes"]]
                notes = str(notes_val).strip() if notes_val else None

            event_type = _detect_event_type(title)

            event = Event(
                season_id=season_id,
                venue_id=venue_id,
                title=title,
                event_type=event_type,
                start_at=start_at,
                end_at=end_at,
                notes=notes,
            )
            db.add(event)
            existing_keys.add(dedup_key)
            report.created += 1

        except Exception as exc:
            logger.exception(f"Erreur import ligne {row_idx}")
            report.errors.append(f"Ligne {row_idx}: {exc}")

    await db.flush()
    return report
